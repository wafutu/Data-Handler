# Python 3.8.5, Openpyxl 3.0.4 
# Name: Michael Tunduli
# Email: wafutu@gmail.com 
# Date : 14th August 2020

""" 
This script is to help in verification 
and data entry of UNSOS Equipment repairs in EMU.
It should compare what has been submitted by the
contractor with the already collected database.
The comparison columns are: 
**** room in terms of (compound / location, building, room)
**** ac serial number
**** ticket number
**** the client contacts
These comparison will be on different tables created from 
different data sources. 
It should show the new serials for the equipment 
and record them in appropriate worksheets
at the same time updating the old worksheet
with the necessary details from the submitted work sheets.

"""
#import the necessary modules required for the script to work
import os
import sys
import  openpyxl
import xlrd
import xlsxwriter

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string, rows_from_range 
from openpyxl.utils.cell import column_index_from_string

workbook1_path = "A:\\0Excel\acdata.xlsx"
workbook2_path = "A:\\0Excel\acraidata.xlsx"

#load the desired xlsx workbook
def open_ac_data_file():
    global wb1
    global ws1
    global ac_data_file
    
    ac_data_file = "A:\\0Excel\\acdata.xlsx"
    print("\nOpening file_1.........")
    wb1 = load_workbook(ac_data_file)
    ws1 = wb1['Sheet1'] # open the specific sheet with data
    #print("\n" + "max_row : " + str(ws1.max_row) + "\nmax_column :" + str(ws1.max_column))
    return(open_ac_data_file)

open_ac_data_file()    

# load the desired second  xlsx workbook that is to be compared
def open_rai_ac_file ():
    global rai_ac_file
    global wb2
    global ws2 
    
    rai_ac_file = "A:\\0Excel\\acraidata.xlsx"
    print("\nOpening file_2..........")
    wb2 = load_workbook(rai_ac_file)
    ws2 = wb2['Sheet1'] # open the specific sheet with data
    #print("\n" + "max_row : " + str(ws2.max_row) + "\nmax_column :" + str(ws2.max_column))
    return(open_rai_ac_file)
    
open_rai_ac_file()

# get a list of all the serials in the database
def ac_serials_list():
    global ac_serials_list
    ac_serials_list = []

    for col in ws1.iter_cols(min_row = 2, min_col = 4, max_col = 4, max_row = 3295+1):
        for cell in col:
            if cell.value != None and cell.value != '':
                ac_serials_list.append(cell.value)
                
    #print("\n----------------\n")        
    #print("THIS IS ALL SERIALS LIST:\n")
    #print("THE NUMBER OF ALL SERIALS FOUND IS :" + str(len(ac_serials_list))+"\n")        
    #print(ac_serials_list)
    return(ac_serials_list)
    
ac_serials_list()

# get a list of ac serials from rai cost summaries submitted
def rai_serials_list():
    global rai_serials_list
    rai_serials_list = []
    
    for col in ws2.iter_cols(min_row = 5, min_col = 3, max_col = 3, max_row = 37):
        for cell in col:
            if cell.value != None and cell.value != '': 
                rai_serials_list.append(cell.value)
    #print("\n----------------")            
    #print("\nTHIS IS RAI SERIALS LIST:\n") 
    #print("THE NUMBER OF RAI SERIALS FOUND IS :" + str(len(rai_serials_list))+"\n")       
    #print(rai_serials_list)
    return(rai_serials_list)
    
rai_serials_list()
        
# compare the serials in the database with those from works done submitted.
# create a list of new serials not in the database
def compare_serials():
    global new_serials_list
    new_serials_list = []  
          
    for i in rai_serials_list:
        if i not in ac_serials_list:
            new_serials_list.append(i)
    #print("\n-----------------\n")          
    #print("THIS ARE NEW SERIALS:")
    #print("THE NUMBER OF NEW SERIALS FOUND IS :" + str(len(new_serials_list))+"\n")
    print(new_serials_list)
    return(new_serials_list)

compare_serials()

# get the row numbers of the new serial numbers

open_rai_ac_file()

#Get the number of rows so that we know how many registrations we need to edit
def new_serial_rows():
    global row_count 
    global new_serial_rows 
    
    row_count = ws2.max_row
    new_serial_rows = []
    for j in range(1, row_count):
        ac_serials_cell = ws2.cell(row = j, column = 3)
        current_ac_serial = ac_serials_cell.value
        if current_ac_serial in new_serials_list:
            current_row =  ac_serials_cell.row 
            new_serial_rows.append(current_row)
    
    print("\n----------\n")        
    print("THIS IS ROW NUMBERS OF NEW SERIALS: ")
    print(new_serial_rows)
    
new_serial_rows()

# copy the record on the rows of the row numbers found above.
#copy the lists of records to excel sheet.

wb2 = load_workbook("A:\\0Excel\\acraidata.xlsx")
ws2 = wb2['Sheet1']
rows = list(ws2.iter_rows(max_col=6, values_only=True))
values = []
row_indices = new_serial_rows
for row_index in row_indices:
    values.append(rows[row_index])

print('\n----------\n' + '\nNEW AC RECORDS LIST :\n' + str(values))

#copy the lists of records to excel sheet. using
# xlrd & xlsxwriter
# open old file
old_path = "A:\\0Excel\\acraidata.xlsx"
old_workbook = xlrd.open_workbook(old_path)
old_worksheet = old_workbook.sheet_by_index(0)

# copy data
all_rows = values

# create new file
new_path = "A:\\0Excel\\template.xlsx"
new_workbook = xlsxwriter.Workbook(new_path)
new_worksheet = new_workbook.add_worksheet()

#populate the new file
for row in range (len(all_rows)):
    for col in range (len(all_rows[2])):
        new_worksheet.write(row, col, all_rows[row][col])
new_workbook.close()

def new_serials_record():
    
    global new_serials_record
    new_serials_record = []

    wb = load_workbook("A:\\0Excel\\acraidata.xlsx")
    ws = wb.worksheets[0]

    for row in ws2.iter_rows(min_row = 5, max_row = 46, min_col=1, max_col=3):
        single_row_values = []
        for cell in row:
            if cell.value in new_serials_list:
                single_row_values.append(cell.value)
        new_serials_record.append(single_row_values)
        
    #print("\n-----------\n")
    #print("THIS IS A RECORD OF NEW AC SERIALS : ")
    #print("THE NUMBER OF NEW AC RECORDS FOUND IS :" + str(len(new_serials_record))+"\n") 
    return(new_serials_record)

new_serials_record()

# paste the records in a new excel file called new_serials list

# get the name of all the rooms that we have and make a list
def all_my_rooms_list():
    global all_my_rooms_list
    all_my_rooms_list = []
    
    for col in ws1.iter_cols(min_row = 5, min_col = 9, max_col = 9, max_row = ws1.max_row +1):
        for cell in col:
            all_my_rooms_list.append(cell.value)
    #print("\n-----------------\n")         
    #print("THIS IS ALL ROOMS LIST:")
    #print("THE NUMBER OF ALL ROOMS FOUND IS :" + str(len(all_my_rooms_list))+"\n")
    #print(all_my_rooms_list)
    return(all_my_rooms_list)

all_my_rooms_list() 

# get the new rooms from the submitted works summary sheets
def rai_rooms_list():
    global rai_rooms_list
    rai_rooms_list = []
    
    for col in ws2.iter_cols(min_row=5, min_col= 2, max_col=2, max_row = 45):
        for cell in col:
            rai_rooms_list.append(cell.value)
    #print("\n-----------------")         
    #print("THIS IS RAI ROOMS LIST:")
    #print("THE NUMBER OF RAI ROOMS FOUND IS :" + str(len(rai_rooms_list))+"\n")
    #print(rai_rooms_list)
    return(rai_rooms_list)

rai_rooms_list()        

# compare the rooms from the current submitted works summary with all the number of rooms that we have
def compare_rooms():
    global new_rooms
    new_rooms = []        
    for i in rai_rooms_list:
        if i not in all_my_rooms_list:
            new_rooms.append(i)
    #print("\n-----------------")        
    #print("THIS ARE NEW ROOMS:")
    #print("THE NUMBER OF NEW ROOMS FOUND IS :" + str(len(new_rooms))+"\n")  
    #print(new_rooms)
    
compare_rooms()

# get the row numbers of the new serial numbers

# copy the record on the rows of the row numbers found above.

# paste the records in a new excel file called new_serials list 

# get the three names of the room as per three rows:
#**** compound or location
#**** building or block
#**** room
def all_3_names_rooms_list():
    global all_3_names_rooms_list
    all_3_names_rooms_list = []

    wb = load_workbook("A:\\0Excel\\acdata.xlsx")
    ws = wb.worksheets[0]
    
    for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):

        single_row_values = []

        for cell in row:
            if cell.value != None and cell.value != "":
                single_row_values.append(cell.value)
        all_3_names_rooms_list.append(single_row_values)

    #print("\n-----------\n")
    #print("THIS IS A LIST OF ALL ROOMS BY 3_NAMES :")
    #print("THE NUMBER OF ALL ROOMS FOUND IS :" + str(len(all_3_names_rooms_list))+"\n") 
    #print(all_3_names_rooms_list)
            
    return all_3_names_rooms_list

all_3_names_rooms_list()

def rai_3_names_rooms_list():
    global rai_3_names_rooms_list
    rai_3_names_rooms_list = []

    wb = load_workbook("A:\\0Excel\\acraidata.xlsx")
    ws = wb.worksheets[0]
    
    for row in ws.iter_rows(min_row=5, max_row=10, min_col=1, max_col=3):
        single_row_values = []
        for cell in row:
            if cell.value != None and cell.value != "":
                single_row_values.append(cell.value)
        rai_3_names_rooms_list.append(single_row_values)

    #print("\n-----------\n")
    #print("THIS IS A LIST OF RAI WORKS ROOMS BY 3_NAMES : ")
    #print("THE NUMBER OF RAI WORKS ROOMS FOUND IS :" + str(len(rai_3_names_rooms_list))+"\n") 
    
    #print(rai_3_names_rooms_list)
            
    return rai_3_names_rooms_list

rai_3_names_rooms_list()

# compare the rooms list and give the new rooms
def new_3_names_room_list():
    global new_3_names_rooms_list
    new_3_names_rooms_list = []

    for item in all_3_names_rooms_list:
        if item not in rai_3_names_rooms_list:
            new_3_names_rooms_list.append(item)
    #print("\n-----------\n")
    #print("THIS IS A LIST OF NEW 3_NAMES FOR ROOMS:")
    #print("THE NUMBER OF ROOMS FOUND IS :" + str(len(new_3_names_rooms_list))+"\n")        
    #print(new_3_names_rooms_list)
    
new_3_names_room_list()

# get the row numbers of the new_3_ names_room_list

# copy the record on the rows of the new_3_ names_room_list above.

# paste the records in a new excel file called new_3_ names_room_list

# copy the specific data to the required places in different worksheets

#getting data from the new rooms and copying to all rooms data:
# opening the source excel file 
filename ="A:\\0Excel\\acraidata.xlsx"
wb1 = load_workbook(filename) 
ws1 = wb1.worksheets[0] 
  
#File to be copied
wb = openpyxl.load_workbook("A:\\0Excel\\acraidata.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name

#File to be pasted into
template = openpyxl.load_workbook("A:\\0Excel\\copied_data.xlsx") #Add file name
temp_sheet = template['Sheet1'] #Add Sheet name

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
         
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
    return(pasteRange)
def createData():
    print("\nProcessing...")
    selectedRange = copyRange(1,5,8,100,sheet) #Change the 4 number values
    pastingRange = pasteRange(1,2,8,55, temp_sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("A:\\0Excel\\acraidata.copy.xlsx")
    print("Range copied and pasted!")
    return(createData)
    
createData()
